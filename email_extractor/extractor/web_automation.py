"""
Web Form Automation
Automates filling web forms from Excel data using Selenium.
"""

import time
import os
from datetime import datetime
from typing import Dict
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook

class WebFormAutomation:
    """Automate filling web forms from Excel data."""
    
    def __init__(self, headless: bool = False, keep_open: bool = False):
        """Initialize the web automation with Chrome driver.
        
        Args:
            headless: If True, runs browser in headless mode
            keep_open: If True, keeps browser open after completion
        """
        self.driver = None
        self.headless = headless
        self.keep_open = keep_open
        self.stats = {
            "processed": 0,
            "skipped": 0,
            "success": 0,
            "failed": 0
        }
        self.excel_path = None
        self.status_updates = {}
    
    def start_browser(self):
        """Start Chrome browser with local driver."""
        print("Starting browser (Chrome)...")
        
        try:
            options = webdriver.ChromeOptions()
            
            if self.headless:
                options.add_argument('--headless')
            
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            
            # Look for chromedriver.exe in project directory
            project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            chromedriver_path = os.path.join(project_dir, "chromedriver.exe")
            
            if not os.path.exists(chromedriver_path):
                print(f"\n✗ chromedriver.exe not found at: {chromedriver_path}")
                print("\nPlease download ChromeDriver:")
                print("1. Check your Chrome version at chrome://settings/help")
                print("2. Download matching driver from https://googlechromelabs.github.io/chrome-for-testing/")
                print("3. Extract chromedriver.exe to the project folder")
                raise FileNotFoundError("chromedriver.exe not found")
            
            service = Service(executable_path=chromedriver_path)
            self.driver = webdriver.Chrome(service=service, options=options)
            self.driver.maximize_window()
            print("✓ Browser started")
            
        except Exception as e:
            print(f"✗ Error starting browser: {e}")
            raise
    
    def stop_browser(self):
        """Stop browser."""
        if self.keep_open:
            print("\n⚠ Browser left open for debugging. Close manually when done.")
            input("Press Enter to close browser and continue...")
        
        if self.driver:
            self.driver.quit()
            print("✓ Browser closed")
    
    def process_excel_file(self, excel_path: str, sheet_name: str):
        """Process an Excel sheet and fill forms for each row."""
        self.excel_path = excel_path
        self.status_updates[sheet_name] = {}
        
        print(f"\n{'='*60}")
        print(f"Processing {sheet_name} Sheet")
        print(f"{'='*60}\n")
        
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        total_rows = len(df)
        print(f"Total rows: {total_rows}\n")
        
        for index, row in df.iterrows():
            row_num = index + 2
            excel_row = index
            
            print(f"\n--- Processing Row {row_num} ---")
            
            take_action = str(row.get("Take Action", "")).strip()
            if not take_action or take_action == "nan":
                print(f"⊘ Skipping - Take Action is empty")
                self.status_updates[sheet_name][excel_row] = "Skipped - No Take Action"
                self.stats["skipped"] += 1
                continue
            
            link = str(row.get("PreMQL review/validation link", "")).strip()
            if not link or link == "nan" or not link.startswith("http"):
                print(f"⊘ Skipping - Invalid or missing link")
                self.status_updates[sheet_name][excel_row] = "Skipped - No Link"
                self.stats["skipped"] += 1
                continue
            
            print(f"Link: {link[:50]}...")
            print(f"Take Action: {take_action}")
            
            if sheet_name == "Validation":
                form_data = self._prepare_validation_data(row)
            else:
                form_data = self._prepare_review_data(row)
            
            success, error_msg = self._fill_form(link, form_data, sheet_name)
            
            if success:
                self.stats["success"] += 1
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.status_updates[sheet_name][excel_row] = f"✓ Success - {timestamp}"
                print(f"✓ Successfully processed row {row_num}")
            else:
                self.stats["failed"] += 1
                self.status_updates[sheet_name][excel_row] = f"✗ Failed - {error_msg}"
                print(f"✗ Failed to process row {row_num}: {error_msg}")
            
            self.stats["processed"] += 1
            time.sleep(1)
        
        self._update_excel_sheet(sheet_name)
        self._print_stats()
    
    def _update_excel_sheet(self, sheet_name: str):
        """Write Form Submission Status back to Excel."""
        if sheet_name not in self.status_updates or not self.status_updates[sheet_name]:
            return
        
        print(f"\nUpdating {sheet_name} sheet in Excel...")
        
        try:
            wb = load_workbook(self.excel_path)
            ws = wb[sheet_name]
            
            status_col = None
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value == "Form Submission Status":
                    status_col = col
                    break
            
            if status_col is None:
                print("  ⊘ Could not find 'Form Submission Status' column")
                return
            
            for excel_row, status in self.status_updates[sheet_name].items():
                excel_write_row = excel_row + 2
                ws.cell(row=excel_write_row, column=status_col, value=status)
            
            wb.save(self.excel_path)
            print(f"  ✓ Updated {len(self.status_updates[sheet_name])} rows in Excel")
            
        except Exception as e:
            print(f"  ✗ Error updating Excel: {e}")
    
    def _prepare_validation_data(self, row: pd.Series) -> Dict[str, str]:
        """Prepare form data from Validation sheet row."""
        data = {
            "take_action": str(row.get("Take Action", "")).strip(),
            "company": str(row.get("Company", "")).strip(),
            "reject_reason": str(row.get("Valid Company → Reject Reason", "")).strip(),
            "invalid_reason": str(row.get("Invalid Company Reason", "")).strip(),
            "additional_info": str(row.get("Additional Scoring Information", "")).strip(),
            "send_to": str(row.get("Send to", "")).strip(),
        }
        
        for key in data:
            if data[key] == "nan":
                data[key] = ""
        
        return data
    
    def _prepare_review_data(self, row: pd.Series) -> Dict[str, str]:
        """Prepare form data from Review sheet row."""
        data = {
            "take_action": str(row.get("Take Action", "")).strip(),
            "company": str(row.get("Company", "")).strip(),
            "reject_reason": str(row.get("Reject Reason", "")).strip(),
            "additional_info": str(row.get("Additional Scoring Information", "")).strip(),
            "send_to": str(row.get("Send to", "")).strip(),
        }
        
        for key in data:
            if data[key] == "nan":
                data[key] = ""
        
        return data
    
    def _fill_form(self, url: str, form_data: Dict[str, str], sheet_type: str) -> tuple:
        """Navigate to URL and fill the form."""
        try:
            self.driver.get(url)
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            time.sleep(2)
            
            if sheet_type == "Validation":
                success, error = self._fill_validation_form(form_data)
            else:
                success, error = self._fill_review_form(form_data)
            
            return success, error
            
        except TimeoutException:
            return False, "Timeout loading page"
        except Exception as e:
            return False, str(e)[:50]
    
    def _fill_validation_form(self, data: Dict[str, str]) -> tuple:
        """Fill validation form fields."""
        try:
            take_action = data["take_action"]
            
            radio_map = {
                "Valid Company → MQL": "valid-acc",
                "Valid Company → Reject": "valid-rej",
                "Invalid Company": "invalid"
            }
            
            if take_action in radio_map:
                radio_id = radio_map[take_action]
                radio = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, radio_id))
                )
                radio.click()
                print(f"  ✓ Selected radio: {take_action}")
                time.sleep(2)  # Wait for form sections to appear
            
            if "MQL" in take_action:
                # Fill Company Name field
                if data["company"]:
                    try:
                        # Wait for the forward section to become visible
                        WebDriverWait(self.driver, 5).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, ".form-section.forward"))
                        )
                        
                        # Find company field by name
                        company_field = WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.NAME, "company"))
                        )
                        company_field.clear()
                        company_field.send_keys(data["company"])
                        print(f"  ✓ Filled Company: {data['company']}")
                    except Exception as e:
                        print(f"  ⊘ Could not fill Company: {e}")
                
                # Fill Additional Scoring Information
                if data["additional_info"]:
                    try:
                        # Wait for the section to be visible
                        WebDriverWait(self.driver, 5).until(
                            EC.visibility_of_element_located((By.ID, "initCallNotes"))
                        )
                        
                        textarea = WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.NAME, "leadDetails1"))
                        )
                        textarea.clear()
                        textarea.send_keys(data["additional_info"])
                        print(f"  ✓ Filled Additional Scoring Information")
                    except Exception as e:
                        print(f"  ⊘ Could not fill Additional Scoring Information: {e}")
                
                # Fill Send to
                if data["send_to"]:
                    try:
                        send_to_field = WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.NAME, "Assignment_compass1"))
                        )
                        send_to_field.clear()
                        send_to_field.send_keys(data["send_to"])
                        print(f"  ✓ Filled Send to: {data['send_to']}")
                    except Exception as e:
                        print(f"  ⊘ Could not fill Send to: {e}")
            
            elif "Reject" in take_action:
                if data["reject_reason"]:
                    try:
                        # Wait for reject section to be visible
                        WebDriverWait(self.driver, 5).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, ".form-section.reject"))
                        )
                        time.sleep(1)
                        self._fill_dropdown("rejectionReason", data["reject_reason"], by_name=True)
                    except Exception as e:
                        print(f"  ⊘ Could not fill rejection reason: {e}")
            
            elif "Invalid" in take_action:
                if data["invalid_reason"]:
                    try:
                        # Wait for invalid section to be visible
                        WebDriverWait(self.driver, 5).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, ".form-section.invalid"))
                        )
                        time.sleep(1)
                        self._fill_dropdown("excludedReason", data["invalid_reason"], by_name=True)
                    except Exception as e:
                        print(f"  ⊘ Could not fill invalid reason: {e}")
            
            self._submit_form()
            return True, ""
            
        except Exception as e:
            return False, str(e)[:100]
    
    def _fill_review_form(self, data: Dict[str, str]) -> tuple:
        """Fill review form fields."""
        try:
            take_action = data["take_action"]
            
            radio_map = {
                "MQL - Send to Sales": "acc",
                "Reject": "rej"
            }
            
            if take_action in radio_map:
                radio_id = radio_map[take_action]
                radio = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, radio_id))
                )
                radio.click()
                print(f"  ✓ Selected radio: {take_action}")
                time.sleep(2)  # Wait for form sections to appear
            
            if "MQL" in take_action or "Send to Sales" in take_action:
                # Fill Additional Scoring Information
                if data["additional_info"]:
                    try:
                        # Wait for the section to be visible
                        WebDriverWait(self.driver, 5).until(
                            EC.visibility_of_element_located((By.ID, "initCallNotes"))
                        )
                        
                        textarea = WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.NAME, "leadDetails1"))
                        )
                        textarea.clear()
                        textarea.send_keys(data["additional_info"])
                        print(f"  ✓ Filled Additional Scoring Information")
                    except Exception as e:
                        print(f"  ⊘ Could not fill Additional Scoring Information: {e}")
                
                # Fill Send to
                if data["send_to"]:
                    try:
                        send_to_field = WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.NAME, "Assignment_compass1"))
                        )
                        send_to_field.clear()
                        send_to_field.send_keys(data["send_to"])
                        print(f"  ✓ Filled Send to: {data['send_to']}")
                    except Exception as e:
                        print(f"  ⊘ Could not fill Send to: {e}")
            
            elif "Reject" in take_action:
                if data["reject_reason"]:
                    try:
                        # Wait for reject section to be visible
                        WebDriverWait(self.driver, 5).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, ".form-section.reject"))
                        )
                        time.sleep(1)
                        self._fill_dropdown("rejectionReason", data["reject_reason"], by_name=True)
                    except Exception as e:
                        print(f"  ⊘ Could not fill rejection reason: {e}")
            
            self._submit_form()
            return True, ""
            
        except Exception as e:
            return False, str(e)[:100]
    
    def _fill_dropdown(self, selector: str, value: str, by_name: bool = False):
        """Fill a dropdown field."""
        if not value:
            return
        
        try:
            if by_name:
                element = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.NAME, selector))
                )
            else:
                element = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
            
            select = Select(element)
            
            # Try exact match
            try:
                select.select_by_visible_text(value)
                print(f"  ✓ Selected dropdown: {value}")
                return
            except:
                pass
            
            # Try case-insensitive match
            for option in select.options:
                option_text = option.text.strip()
                if option_text.lower() == value.lower():
                    select.select_by_visible_text(option_text)
                    print(f"  ✓ Selected dropdown: {option_text}")
                    return
            
            # Try partial match
            for option in select.options:
                option_text = option.text.strip()
                if value.lower() in option_text.lower() or option_text.lower() in value.lower():
                    select.select_by_visible_text(option_text)
                    print(f"  ✓ Selected dropdown (partial match): {option_text}")
                    return
            
            print(f"  ⊘ Could not find '{value}' in dropdown. Available options:")
            for option in select.options:
                print(f"      - '{option.text.strip()}'")
            
            raise Exception(f"Value '{value}' not found in dropdown")
            
        except Exception as e:
            print(f"  ⊘ Could not fill dropdown: {e}")
            raise
    
    def _submit_form(self):
        """Submit the form."""
        try:
            submit_button = self.driver.find_element(By.ID, "submitBtn")
            submit_button.click()
            time.sleep(3)
            print(f"  ✓ Form submitted")
        except Exception as e:
            print(f"  ⊘ Could not submit form: {e}")
            try:
                submit_button = self.driver.find_element(By.CSS_SELECTOR, "input[type='submit']")
                submit_button.click()
                time.sleep(3)
                print(f"  ✓ Form submitted (fallback)")
            except:
                raise
    
    def _print_stats(self):
        """Print processing statistics."""
        print(f"\n{'='*60}")
        print("PROCESSING SUMMARY")
        print(f"{'='*60}")
        print(f"Total processed:  {self.stats['processed']}")
        print(f"Successful:       {self.stats['success']}")
        print(f"Failed:           {self.stats['failed']}")
        print(f"Skipped:          {self.stats['skipped']}")
        print(f"{'='*60}\n")