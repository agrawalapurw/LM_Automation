"""
Excel Writer
Writes DataFrames to Excel with formatting and dropdowns.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule

# Color scheme
FILTER_COLUMNS_COLOR = "FFC7CE"      # Light Red
USER_INPUT_COLOR = "FFEB9C"          # Light Orange
STATUS_TRACKING_COLOR = "C6EFCE"     # Light Green

# Column groups for Validation sheet
VALIDATION_FILTER_COLUMNS = [
    "Has Contact Sales Form",
    "Company Domain Validation",
    "Validation Status",
    "Status"
]

VALIDATION_INPUT_COLUMNS = [
    "Take Action",
    "Valid Company → Reject Reason",
    "Invalid Company Reason",
    "Additional Scoring Information",
    "Send to",
    "Move to Folder"
]

VALIDATION_STATUS_COLUMNS = [
    "Action Taken",
    "Form Submission Status",
    "Email Move Status"
]

# Column groups for Review sheet
REVIEW_FILTER_COLUMNS = [
    "Has Contact Sales Form",
    "Company Domain Validation",
    "Validation Status",
    "Status"
]

REVIEW_INPUT_COLUMNS = [
    "Take Action",
    "Reject Reason",
    "Additional Scoring Information",
    "Send to",
    "Move to Folder"
]

REVIEW_STATUS_COLUMNS = [
    "Action Taken",
    "Form Submission Status",
    "Email Move Status"
]

# Dropdown options
TAKE_ACTION_VALIDATION = [
    "Valid Company → MQL",
    "Valid Company → Reject",
    "Invalid Company"
]

VALID_REJECT_REASONS_VALIDATION = [
    "Not able to contact",
    "Not a Disti lead",
    "Contacted - no general potential",
    "Contacted - no current potential",
    "Not contacted - no general potential",
    "Insufficient lead profile information",
    "Lead already known",
    "University Contact",
    "Distribution Partner"
]

INVALID_COMPANY_REASONS = [
    "University Contact",
    "Distribution Partner",
    "Company - with no potential",
    "Agency",
    "Free-mailer with no potential",
    "Competitor"
]

TAKE_ACTION_REVIEW = [
    "MQL - Send to Sales",
    "Reject"
]

REJECT_REASONS_REVIEW = [
    "Not able to contact",
    "Not a Disti lead",
    "Contacted - no general potential",
    "Contacted - no current potential",
    "Not contacted - no general potential",
    "Insufficient lead profile information",
    "Lead already known",
    "University Contact",
    "Distribution Partner"
]

MOVE_TO_FOLDER_OPTIONS = [
    "Arrow",
    "EBV/Avnet",
    "Future",
    "Non-EBV Leads",
    "Other Distribution Partners",
    "Rutronik",
    "Rejected Marketing"
]

FOLDER_COLORS = {
    "Arrow": "FFE699",
    "EBV/Avnet": "B4C7E7",
    "Future": "C5E0B4",
    "Non-EBV Leads": "F4B084",
    "Other Distribution Partners": "D9D2E9",
    "Rutronik": "FFF2CC",
    "Rejected Marketing": "F8CBAD"
}

class ExcelWriter:
    """Write DataFrames to Excel with separate sheets and dropdowns."""
    
    def write_workbook(self, df_validation: pd.DataFrame, df_review: pd.DataFrame, filepath: str):
        """Write two DataFrames to Excel with separate sheets."""
        wb = Workbook()
        
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        if not df_validation.empty:
            ws_val = wb.create_sheet("Validation")
            self._write_validation_sheet(ws_val, df_validation)
        
        if not df_review.empty:
            ws_rev = wb.create_sheet("Review")
            self._write_review_sheet(ws_rev, df_review)
        
        wb.save(filepath)
    
    def _write_validation_sheet(self, worksheet, df: pd.DataFrame):
        """Write Validation sheet with dropdowns."""
        all_workflow = VALIDATION_FILTER_COLUMNS + VALIDATION_INPUT_COLUMNS + VALIDATION_STATUS_COLUMNS
        
        for col in all_workflow:
            if col not in df.columns:
                df[col] = ""
        
        if "Reject Reason" in df.columns:
            df = df.drop(columns=["Reject Reason"])
        
        data_columns = [col for col in df.columns if col not in all_workflow]
        ordered_columns = (data_columns + 
                          VALIDATION_FILTER_COLUMNS + 
                          VALIDATION_INPUT_COLUMNS + 
                          VALIDATION_STATUS_COLUMNS)
        df = df[ordered_columns]
        
        headers = list(df.columns)
        worksheet.append(headers)
        
        for _, row in df.iterrows():
            worksheet.append([row[col] for col in headers])
        
        self._color_headers(worksheet, headers, "Validation")
        
        take_action_idx = headers.index("Take Action") + 1
        valid_reject_idx = headers.index("Valid Company → Reject Reason") + 1
        invalid_idx = headers.index("Invalid Company Reason") + 1
        move_to_folder_idx = headers.index("Move to Folder") + 1
        
        last_row = worksheet.max_row
        total_cols = len(headers)
        
        # Add dropdowns
        self._add_dropdown(worksheet, take_action_idx, last_row, TAKE_ACTION_VALIDATION)
        self._add_dropdown(worksheet, valid_reject_idx, last_row, VALID_REJECT_REASONS_VALIDATION)
        self._add_dropdown(worksheet, invalid_idx, last_row, INVALID_COMPANY_REASONS)
        self._add_dropdown(worksheet, move_to_folder_idx, last_row, MOVE_TO_FOLDER_OPTIONS)
        
        # Add conditional formatting
        self._add_conditional_formatting_validation(worksheet, headers, last_row)
        
        # Add row coloring
        self._add_row_coloring(worksheet, move_to_folder_idx, last_row, total_cols)
        
        # Make links clickable
        self._make_links_clickable(worksheet, "PreMQL review/validation link")
        self._make_links_clickable(worksheet, "Eloqua Profiler")
        self._make_links_clickable(worksheet, "URL Of Form")
        
        worksheet.freeze_panes = "A2"
        self._adjust_column_widths(worksheet)
    
    def _write_review_sheet(self, worksheet, df: pd.DataFrame):
        """Write Review sheet with dropdowns."""
        all_workflow = REVIEW_FILTER_COLUMNS + REVIEW_INPUT_COLUMNS + REVIEW_STATUS_COLUMNS
        
        for col in all_workflow:
            if col not in df.columns:
                df[col] = ""
        
        validation_specific = [
            "Valid Company → Reject Reason",
            "Invalid Company Reason"
        ]
        df = df.drop(columns=[col for col in validation_specific if col in df.columns])
        
        data_columns = [col for col in df.columns if col not in all_workflow]
        ordered_columns = (data_columns + 
                          REVIEW_FILTER_COLUMNS + 
                          REVIEW_INPUT_COLUMNS + 
                          REVIEW_STATUS_COLUMNS)
        df = df[ordered_columns]
        
        headers = list(df.columns)
        worksheet.append(headers)
        
        for _, row in df.iterrows():
            worksheet.append([row[col] for col in headers])
        
        self._color_headers(worksheet, headers, "Review")
        
        take_action_idx = headers.index("Take Action") + 1
        reject_reason_idx = headers.index("Reject Reason") + 1
        move_to_folder_idx = headers.index("Move to Folder") + 1
        
        last_row = worksheet.max_row
        total_cols = len(headers)
        
        # Add dropdowns
        self._add_dropdown(worksheet, take_action_idx, last_row, TAKE_ACTION_REVIEW)
        self._add_dropdown(worksheet, reject_reason_idx, last_row, REJECT_REASONS_REVIEW)
        self._add_dropdown(worksheet, move_to_folder_idx, last_row, MOVE_TO_FOLDER_OPTIONS)
        
        # Add conditional formatting
        self._add_conditional_formatting_review(worksheet, headers, last_row)
        
        # Add row coloring
        self._add_row_coloring(worksheet, move_to_folder_idx, last_row, total_cols)
        
        # Make links clickable
        self._make_links_clickable(worksheet, "PreMQL review/validation link")
        self._make_links_clickable(worksheet, "Eloqua Profiler")
        self._make_links_clickable(worksheet, "URL Of Form")
        
        worksheet.freeze_panes = "A2"
        self._adjust_column_widths(worksheet)
    
    def _add_dropdown(self, worksheet, col_idx: int, last_row: int, options: list):
        """Add dropdown validation to a column."""
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=f'"' + ','.join(options) + '"',
            allow_blank=True
        )
        worksheet.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")
    
    def _add_row_coloring(self, worksheet, move_to_folder_col: int, last_row: int, total_cols: int):
        """Add conditional formatting to color rows based on Move to Folder."""
        move_to_folder_letter = get_column_letter(move_to_folder_col)
        
        for folder, color in FOLDER_COLORS.items():
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            first_col = get_column_letter(1)
            last_col = get_column_letter(total_cols)
            range_to_format = f"{first_col}2:{last_col}{last_row}"
            
            formula = f'${move_to_folder_letter}2="{folder}"'
            rule = FormulaRule(formula=[formula], stopIfTrue=False, fill=fill)
            
            worksheet.conditional_formatting.add(range_to_format, rule)
    
    def _color_headers(self, worksheet, headers: list, sheet_type: str):
        """Color-code headers based on column groups."""
        if sheet_type == "Validation":
            filter_cols = VALIDATION_FILTER_COLUMNS
            input_cols = VALIDATION_INPUT_COLUMNS
            status_cols = VALIDATION_STATUS_COLUMNS
        else:
            filter_cols = REVIEW_FILTER_COLUMNS
            input_cols = REVIEW_INPUT_COLUMNS
            status_cols = REVIEW_STATUS_COLUMNS
        
        filter_fill = PatternFill(start_color=FILTER_COLUMNS_COLOR, end_color=FILTER_COLUMNS_COLOR, fill_type="solid")
        input_fill = PatternFill(start_color=USER_INPUT_COLOR, end_color=USER_INPUT_COLOR, fill_type="solid")
        status_fill = PatternFill(start_color=STATUS_TRACKING_COLOR, end_color=STATUS_TRACKING_COLOR, fill_type="solid")
        default_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        header_font_white = Font(bold=True, color="FFFFFF")
        header_font_black = Font(bold=True, color="000000")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, col_name in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.alignment = header_alignment
            
            if col_name in filter_cols:
                cell.fill = filter_fill
                cell.font = header_font_black
            elif col_name in input_cols:
                cell.fill = input_fill
                cell.font = header_font_black
            elif col_name in status_cols:
                cell.fill = status_fill
                cell.font = header_font_black
            else:
                cell.fill = default_fill
                cell.font = header_font_white
        
        worksheet.row_dimensions[1].height = 35
    
    def _make_links_clickable(self, worksheet, column_name: str):
        """Make URLs in a specific column clickable."""
        link_col = None
        
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value == column_name:
                link_col = col
                break
        
        if link_col is None:
            return
        
        link_font = Font(color="0000FF", underline="single")
        
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=link_col)
            url = cell.value
            
            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.value = url
                cell.font = link_font
                cell.style = "Hyperlink"
    
    def _add_conditional_formatting_validation(self, worksheet, headers: list, last_row: int):
        """Add conditional formatting for Validation sheet."""
        highlight_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        highlight_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        take_action_idx = headers.index("Take Action") + 1
        valid_reject_idx = headers.index("Valid Company → Reject Reason") + 1
        invalid_idx = headers.index("Invalid Company Reason") + 1
        additional_idx = headers.index("Additional Scoring Information") + 1
        send_to_idx = headers.index("Send to") + 1
        
        take_action_letter = get_column_letter(take_action_idx)
        valid_reject_letter = get_column_letter(valid_reject_idx)
        invalid_letter = get_column_letter(invalid_idx)
        additional_letter = get_column_letter(additional_idx)
        send_to_letter = get_column_letter(send_to_idx)
        
        for row in range(2, last_row + 1):
            # Highlight Valid Company → Reject Reason
            rule1 = FormulaRule(
                formula=[f'${take_action_letter}{row}="Valid Company → Reject"'],
                stopIfTrue=False,
                fill=highlight_yellow
            )
            worksheet.conditional_formatting.add(f"{valid_reject_letter}{row}", rule1)
            
            # Highlight Invalid Company Reason
            rule2 = FormulaRule(
                formula=[f'${take_action_letter}{row}="Invalid Company"'],
                stopIfTrue=False,
                fill=highlight_yellow
            )
            worksheet.conditional_formatting.add(f"{invalid_letter}{row}", rule2)
            
            # Highlight Additional Scoring Info and Send to
            rule3 = FormulaRule(
                formula=[f'${take_action_letter}{row}="Valid Company → MQL"'],
                stopIfTrue=False,
                fill=highlight_green
            )
            worksheet.conditional_formatting.add(f"{additional_letter}{row}", rule3)
            worksheet.conditional_formatting.add(f"{send_to_letter}{row}", rule3)
    
    def _add_conditional_formatting_review(self, worksheet, headers: list, last_row: int):
        """Add conditional formatting for Review sheet."""
        highlight_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        highlight_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        take_action_idx = headers.index("Take Action") + 1
        reject_reason_idx = headers.index("Reject Reason") + 1
        additional_idx = headers.index("Additional Scoring Information") + 1
        send_to_idx = headers.index("Send to") + 1
        
        take_action_letter = get_column_letter(take_action_idx)
        reject_reason_letter = get_column_letter(reject_reason_idx)
        additional_letter = get_column_letter(additional_idx)
        send_to_letter = get_column_letter(send_to_idx)
        
        for row in range(2, last_row + 1):
            # Highlight Reject Reason
            rule1 = FormulaRule(
                formula=[f'${take_action_letter}{row}="Reject"'],
                stopIfTrue=False,
                fill=highlight_yellow
            )
            worksheet.conditional_formatting.add(f"{reject_reason_letter}{row}", rule1)
            
            # Highlight Additional Scoring Info and Send to
            rule2 = FormulaRule(
                formula=[f'${take_action_letter}{row}="MQL - Send to Sales"'],
                stopIfTrue=False,
                fill=highlight_green
            )
            worksheet.conditional_formatting.add(f"{additional_letter}{row}", rule2)
            worksheet.conditional_formatting.add(f"{send_to_letter}{row}", rule2)
    
    def _adjust_column_widths(self, worksheet):
        """Auto-adjust column widths."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width