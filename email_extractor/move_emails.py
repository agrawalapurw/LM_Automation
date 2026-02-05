"""
Move Emails Tool
Moves emails to folders based on Excel 'Move to Folder' column.
"""
from __future__ import annotations
import argparse
import sys
from pathlib import Path
from typing import Dict, Optional
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
import win32com.client

# Canonical names and aliases
FOLDER_ALIASES = {
    "arrow": ["arrow"],
    "future": ["future"],
    "rutronik": ["rutronik"],
    "ebv/avnet": ["ebv", "avnet", "ebv/avnet", "ebv avnet"],
    "non-ebv leads": ["non ebv", "non-ebv", "non ebv leads", "non ebv lead"],
    "other distribution partners": ["other distribution partners", "other distribution", "other dist"],
    "rejected marketing": ["rejected marketing", "rejected", "marketing rejected"],
}

def normalize_folder_name(name: str) -> str:
    n = name.strip().lower()
    for canonical, aliases in FOLDER_ALIASES.items():
        if n == canonical or any(n == a for a in aliases):
            return canonical
    return n

class SmartEmailMover:
    """Smart email mover that finds folders automatically."""

    def __init__(self):
        self.outlook = win32com.client.Dispatch("Outlook.Application")
        self.namespace = self.outlook.GetNamespace("MAPI")
        self.folder_cache: Dict[str, object] = {}
        self.stats = {
            "moved": 0,
            "failed": 0,
            "skipped": 0,
            "folder_not_found": 0,
        }

    def find_folder_recursive(self, root_folder, target_name: str, max_depth: int = 10, current_depth: int = 0):
        if current_depth > max_depth:
            return None
        try:
            for folder in root_folder.Folders:
                name = (folder.Name or "").strip()
                n_lower = name.lower()
                t_lower = target_name.lower()
                if n_lower == t_lower or t_lower in n_lower:
                    return folder
                result = self.find_folder_recursive(folder, target_name, max_depth, current_depth + 1)
                if result:
                    return result
        except Exception:
            pass
        return None

    def find_folder_in_all_stores(self, target_name: str):
        cache_key = target_name.lower()
        if cache_key in self.folder_cache:
            return self.folder_cache[cache_key]

        # Build search names from aliases
        canonical = normalize_folder_name(target_name)
        search_names = set([target_name, canonical])
        # Add all aliases for canonical
        for canon, aliases in FOLDER_ALIASES.items():
            if canonical == canon:
                search_names.update(aliases)

        # Search in Stores
        try:
            for i in range(self.namespace.Stores.Count):
                store = self.namespace.Stores.Item(i + 1)
                root = store.GetRootFolder()
                for name in search_names:
                    found = self.find_folder_recursive(root, name)
                    if found:
                        self.folder_cache[cache_key] = found
                        print(f"  ✓ Found '{target_name}' at: {found.FolderPath}")
                        return found
        except Exception:
            pass

        # Fallback for older Outlook versions
        try:
            for i in range(self.namespace.Folders.Count):
                root = self.namespace.Folders.Item(i + 1)
                for name in search_names:
                    found = self.find_folder_recursive(root, name)
                    if found:
                        self.folder_cache[cache_key] = found
                        print(f"  ✓ Found '{target_name}' at: {found.FolderPath}")
                        return found
        except Exception:
            pass

        print(f"  ✗ Folder not found: {target_name}")
        return None

    def get_item_by_entry_id(self, entry_id: str):
        if not entry_id:
            return None
        try:
            return self.namespace.GetItemFromID(entry_id)
        except Exception:
            return None

    def restrict_items_by_time(self, folder, dt: datetime, window_minutes: int = 15):
        items = folder.Items
        items.Sort("[ReceivedTime]", True)
        fmt = "%m/%d/%Y %I:%M %p"
        start = (dt - timedelta(minutes=window_minutes)).strftime(fmt)
        end = (dt + timedelta(minutes=window_minutes)).strftime(fmt)
        query = f"[ReceivedTime] >= '{start}' AND [ReceivedTime] <= '{end}'"
        try:
            return items.Restrict(query)
        except Exception:
            items.IncludeRecurrences = False
            return items.Restrict(query)

    def normalize_subject(self, s: str) -> str:
        s = (s or "").strip().lower()
        for prefix in ("re:", "fw:", "fwd:"):
            if s.startswith(prefix):
                s = s[len(prefix):].strip()
        return s

    def parse_received_time(self, value: str) -> Optional[datetime]:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except Exception:
                continue
        return None

    def find_email(self, source_folder, subject: str, received_time: str, entry_id: Optional[str]) -> Optional[object]:
        # 1) Prefer EntryID
        if entry_id:
            item = self.get_item_by_entry_id(entry_id)
            if item:
                return item

        # 2) Restrict by time window and match normalized subject
        dt = self.parse_received_time(received_time)
        if dt:
            try:
                restricted = self.restrict_items_by_time(source_folder, dt, window_minutes=15)
                target_subject = self.normalize_subject(subject)
                for i in range(1, restricted.Count + 1):
                    try:
                        itm = restricted.Item(i)
                        if getattr(itm, "Class", None) != 43:
                            continue
                        itm_subject = self.normalize_subject(getattr(itm, "Subject", "") or "")
                        if itm_subject == target_subject:
                            return itm
                    except Exception:
                        continue
            except Exception:
                pass

        # 3) Fallback: limited scan
        try:
            items = source_folder.Items
            items.Sort("[ReceivedTime]", True)
            target_subject = self.normalize_subject(subject)
            for i in range(1, min(items.Count + 1, 1000)):
                try:
                    itm = items.Item(i)
                    if getattr(itm, "Class", None) != 43:
                        continue
                    itm_subject = self.normalize_subject(getattr(itm, "Subject", "") or "")
                    if itm_subject == target_subject:
                        return itm
                except Exception:
                    continue
        except Exception:
            pass

        return None

    def move_email(self, email_item, target_folder) -> bool:
        try:
            email_item.Move(target_folder)
            return True
        except Exception as e:
            print(f"  ✗ Error moving email: {e}")
            return False

    def process_excel_file(self, excel_path: Path, source_folder):
        print("\n" + "=" * 60)
        print("EMAIL MOVING AUTOMATION")
        print("=" * 60 + "\n")

        for sheet_name in ("Validation", "Review"):
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
            except Exception:
                print(f"⊘ Sheet '{sheet_name}' not found, skipping...")
                continue

            print(f"\nProcessing {sheet_name} Sheet ({len(df)} rows)")
            print("-" * 60)
            self._process_sheet(df, sheet_name, source_folder, excel_path)

        self._print_summary()

    def _process_sheet(self, df: pd.DataFrame, sheet_name: str, source_folder, excel_path: Path):
        status_updates: Dict[int, str] = {}

        for index, row in df.iterrows():
            row_num = index + 2  # Excel row number
            move_to = row.get("Move to Folder")

            if pd.isna(move_to) or not str(move_to).strip():
                status_updates[index] = "Not processed - No folder selected"
                self.stats["skipped"] += 1
                continue

            subject = str(row.get("Subject", "")).strip()
            received_time = str(row.get("ReceivedTime", "")).strip()
            entry_id = str(row.get("EntryID", "")).strip() if "EntryID" in df.columns else ""

            if not subject or not received_time:
                print(f"Row {row_num}: ⊘ Skipped - Missing subject or time")
                status_updates[index] = "Skipped - Missing identifiers"
                self.stats["skipped"] += 1
                continue

            target_name = normalize_folder_name(str(move_to))
            print(f"\nRow {row_num}: {subject[:40]}... → {target_name}")

            target_folder = self.find_folder_in_all_stores(target_name)
            if not target_folder:
                status_updates[index] = f"Failed - Folder '{move_to}' not found"
                self.stats["folder_not_found"] += 1
                continue

            email_item = self.find_email(source_folder, subject, received_time, entry_id)
            if not email_item:
                print("  ✗ Email not found in source folder")
                status_updates[index] = "Failed - Email not found in source folder"
                self.stats["failed"] += 1
                continue

            if self.move_email(email_item, target_folder):
                print("  ✓ Moved successfully")
                status_updates[index] = f"✓ Moved to {move_to}"
                self.stats["moved"] += 1
            else:
                status_updates[index] = f"Failed - Could not move to {move_to}"
                self.stats["failed"] += 1

        self._update_excel_status(excel_path, sheet_name, status_updates)

    def _update_excel_status(self, excel_path: Path, sheet_name: str, status_updates: Dict[int, str]):
        if not status_updates:
            return

        print("\n  Updating Excel with move status...")
        try:
            wb = load_workbook(excel_path)
            ws = wb[sheet_name]

            status_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "Email Move Status":
                    status_col = col
                    break

            if status_col is None:
                status_col = ws.max_column + 1
                ws.cell(row=1, column=status_col, value="Email Move Status")

            for excel_row, status in status_updates.items():
                ws.cell(row=excel_row + 2, column=status_col, value=status)

            wb.save(excel_path)
            print(f"  ✓ Updated {len(status_updates)} rows in Excel")
        except Exception as e:
            print(f"  ✗ Error updating Excel: {e}")

    def _print_summary(self):
        print("\n" + "=" * 60)
        print("EMAIL MOVING SUMMARY")
        print("=" * 60)
        print(f"Successfully moved:    {self.stats['moved']}")
        print(f"Failed to move:        {self.stats['failed']}")
        print(f"Folder not found:      {self.stats['folder_not_found']}")
        print(f"Skipped (no folder):   {self.stats['skipped']}")
        print("=" * 60 + "\n")


def select_source_folder(namespace):
    """Interactive folder selection (unchanged API)."""
    stores = []
    try:
        for i in range(namespace.Stores.Count):
            stores.append(namespace.Stores.Item(i + 1).DisplayName)
    except Exception:
        for i in range(namespace.Folders.Count):
            stores.append(namespace.Folders.Item(i + 1).Name)

    if not stores:
        raise RuntimeError("No Outlook stores found")

    print("\nAvailable stores:")
    for i, store in enumerate(stores, 1):
        print(f"  [{i}] {store}")

    while True:
        try:
            choice = int(input("Select store: ").strip())
            if 1 <= choice <= len(stores):
                store_name = stores[choice - 1]
                break
        except Exception:
            pass
        print("Invalid selection.")

    root = None
    try:
        for i in range(namespace.Stores.Count):
            store = namespace.Stores.Item(i + 1)
            if store.DisplayName == store_name:
                root = store.GetRootFolder()
                break
    except Exception:
        pass

    if not root:
        try:
            for i in range(namespace.Folders.Count):
                folder = namespace.Folders.Item(i + 1)
                if folder.Name == store_name:
                    root = folder
                    break
        except Exception:
            pass

    if not root:
        raise RuntimeError(f"Could not access store: {store_name}")

    path = [root]
    while True:
        current = path[-1]
        print(f"\nCurrent: {current.Name}")
        try:
            subfolders = list(current.Folders)
        except Exception:
            subfolders = []

        if subfolders:
            print("Subfolders:")
            for i, folder in enumerate(subfolders, 1):
                print(f"  [{i}] {folder.Name}")

        cmd = input("Enter number, 's' to select, 'u' for up, 'q' to quit: ").strip().lower()
        if cmd == 's':
            return current
        elif cmd == 'u':
            if len(path) > 1:
                path.pop()
        elif cmd == 'q':
            sys.exit(0)
        else:
            try:
                idx = int(cmd) - 1
                if 0 <= idx < len(subfolders):
                    path.append(subfolders[idx])
            except Exception:
                print("Invalid input.")


def main():
    print("=" * 60)
    print("EMAIL MOVING TOOL")
    print("=" * 60)
    print("\nThis tool will:")
    print("1. Read 'Move to Folder' column from Excel")
    print("2. Prefer exact message by EntryID (if present)")
    print("3. Find target folders automatically")
    print("4. Move emails and update Excel status")
    print("=" * 60)

    parser = argparse.ArgumentParser(description="Move emails based on Excel decisions.")
    parser.add_argument("--excel", type=str, help="Path to Excel file in output/")
    parser.add_argument("--source", type=str, help="Search name for source folder (optional)")
    args = parser.parse_args()

    output_dir = Path(__file__).with_name("output")
    if not output_dir.exists():
        print(f"\n✗ Output directory not found: {output_dir}")
        return

    excel_path: Optional[Path] = None
    if args.excel:
        excel_path = Path(args.excel)
        if not excel_path.exists():
            print(f"\n✗ Excel not found: {excel_path}")
            return
    else:
        excel_files = [p for p in output_dir.iterdir() if p.suffix == ".xlsx" and not p.name.startswith("~$")]
        if not excel_files:
            print(f"\n✗ No Excel files found in {output_dir}")
            return
        print("\nAvailable Excel files:")
        for i, p in enumerate(excel_files, 1):
            print(f"  [{i}] {p.name}")
        while True:
            try:
                choice = int(input("\nSelect file number: ").strip())
                if 1 <= choice <= len(excel_files):
                    excel_path = excel_files[choice - 1]
                    break
            except Exception:
                pass
            print("Invalid selection.")

    print("\nSelect the folder where Pre-MQL emails are currently located:")
    mover = SmartEmailMover()

    if args.source:
        # Try to auto-find source folder if a name is provided
        source_folder = mover.find_folder_in_all_stores(args.source)
        if not source_folder:
            print(f"✗ Could not auto-find source folder '{args.source}'. Falling back to interactive selection.")
            source_folder = select_source_folder(mover.namespace)
    else:
        source_folder = select_source_folder(mover.namespace)

    print(f"\n✓ Source folder: {source_folder.FolderPath}")

    print("\n" + "=" * 60)
    print("READY TO MOVE EMAILS")
    print("=" * 60)
    print(f"Excel file: {excel_path.name}")
    print(f"Source folder: {source_folder.Name}")
    print("=" * 60)

    confirm = input("\nProceed? [Y/n]: ").strip().lower()
    if confirm and confirm not in ("y", "yes"):
        print("Cancelled.")
        return

    print("\nStarting email moving process...\n")
    mover.process_excel_file(excel_path, source_folder)
    print("\n✓ Email moving completed!")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()